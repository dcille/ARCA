"""Parse and validate Excel file for custom control import."""
import io
import logging

import openpyxl

from scanner.registry import get_default_registry
from scanner.registry.models import ProviderType, Severity, Category

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = {"check_id", "title", "severity", "provider"}

VALID_PROVIDERS = {p.value for p in ProviderType}
VALID_SEVERITIES = {s.value for s in Severity}
VALID_CATEGORIES = {c.value for c in Category}


def parse_and_validate_excel(content: bytes) -> dict:
    """Parse Excel bytes and validate each row against the registry.

    Returns: {valid: [...], errors: [...], warnings: [...], total_rows: int}
    """
    registry = get_default_registry()

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb["Controls"] if "Controls" in wb.sheetnames else wb.active

    # Read headers from row 1
    headers = []
    for cell in ws[1]:
        val = str(cell.value or "").strip().lower()
        headers.append(val)

    if not REQUIRED_COLUMNS.issubset(set(headers)):
        missing = REQUIRED_COLUMNS - set(headers)
        return {
            "valid": [],
            "errors": [{"row": 1, "message": f"Missing required columns: {', '.join(sorted(missing))}"}],
            "warnings": [],
            "total_rows": 0,
        }

    valid = []
    errors = []
    warnings = []
    seen_ids: set = set()
    row_num = 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        row_data = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_data[headers[i]] = str(val).strip() if val is not None else ""

        # Skip empty rows
        check_id = row_data.get("check_id", "")
        if not check_id or check_id == "None":
            continue

        row_errors = []
        row_warnings = []

        # Required fields
        if not check_id:
            row_errors.append("check_id is required")
        if not row_data.get("title"):
            row_errors.append("title is required")

        # Uniqueness
        if check_id in seen_ids:
            row_errors.append(f"Duplicate check_id '{check_id}'")
        seen_ids.add(check_id)

        # Check if already in registry
        if registry.has_check(check_id):
            row_errors.append(f"check_id '{check_id}' already exists in the registry")

        # Validate provider
        provider = row_data.get("provider", "").lower()
        if provider and provider not in VALID_PROVIDERS:
            row_errors.append(f"Invalid provider '{provider}'. Must be one of: {sorted(VALID_PROVIDERS)}")
        elif not provider:
            row_errors.append("provider is required")

        # Validate severity
        severity = row_data.get("severity", "medium").lower()
        if severity not in VALID_SEVERITIES:
            row_errors.append(f"Invalid severity '{severity}'. Must be one of: {sorted(VALID_SEVERITIES)}")

        # Validate category
        category = row_data.get("category", "Compliance")
        if category and category not in VALID_CATEGORIES:
            row_warnings.append(f"Category '{category}' not in standard list, defaulting to 'Compliance'")
            category = "Compliance"

        # Validate scanner check IDs
        scanner_ids_raw = row_data.get("mapped_scanner_check_ids", "")
        scanner_ids = [s.strip() for s in scanner_ids_raw.split(",") if s.strip()] if scanner_ids_raw else []

        for sid in scanner_ids:
            if not registry.has_scanner_id(sid):
                row_warnings.append(f"scanner_check_id '{sid}' not found in registry")

        valid_scanner_ids = [s for s in scanner_ids if registry.has_scanner_id(s)]
        assessment_type = "automated" if valid_scanner_ids else "manual"

        # Parse tags
        tags_raw = row_data.get("tags", "")
        tags = [t.strip() for t in tags_raw.split(",") if t.strip()] if tags_raw else []

        # Parse references
        refs_raw = row_data.get("references", "")
        references = [r.strip() for r in refs_raw.split(",") if r.strip()] if refs_raw else []

        if row_errors:
            errors.append({"row": row_num, "message": "; ".join(row_errors), "check_id": check_id})
        else:
            entry = {
                "check_id": check_id,
                "title": row_data.get("title", ""),
                "description": row_data.get("description", ""),
                "severity": severity,
                "provider": provider,
                "service": row_data.get("service", "custom") or "custom",
                "category": category,
                "risks": row_data.get("risks", ""),
                "remediation": row_data.get("remediation", ""),
                "scanner_check_ids": scanner_ids,
                "tags": tags,
                "references": references,
            }
            valid.append(entry)

        if row_warnings:
            for w in row_warnings:
                warnings.append({"row": row_num, "message": w, "check_id": check_id})

    wb.close()

    return {
        "valid": valid,
        "errors": errors,
        "warnings": warnings,
        "total_rows": row_num - 1,
    }
