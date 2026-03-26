"""Generate Excel template for custom control import, populated with registry data."""
import io
from typing import BinaryIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scanner.registry import get_default_registry


HEADER_FILL = PatternFill(start_color="012169", end_color="012169", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

CONTROL_HEADERS = [
    ("check_id", 25, "Unique ID for the control (e.g., custom_sec_001)"),
    ("title", 50, "Title of the control"),
    ("description", 60, "Detailed description"),
    ("severity", 15, "critical | high | medium | low | informational"),
    ("provider", 15, "aws | azure | gcp | oci | etc."),
    ("service", 20, "Service name (default: custom)"),
    ("category", 20, "Category (Identity, Encryption, Networking, etc.)"),
    ("assessment_type", 18, "Will be auto-determined from scanner mappings"),
    ("risks", 40, "Risk description if control is not met"),
    ("remediation", 40, "Remediation guidance"),
    ("mapped_scanner_check_ids", 40, "Comma-separated scanner check IDs from the registry"),
    ("tags", 30, "Comma-separated tags"),
    ("references", 40, "Comma-separated reference URLs"),
]


def generate_template() -> BinaryIO:
    """Generate and return an Excel template as a BytesIO stream."""
    registry = get_default_registry()
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Controls (for the user to fill) ----
    ws = wb.active
    ws.title = "Controls"

    for col_idx, (header, width, tooltip) in enumerate(CONTROL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        if tooltip:
            cell.comment = openpyxl.comments.Comment(tooltip, "ARCA")

    # Add example row
    example = [
        "custom_example_001",
        "Ensure CloudTrail is enabled in all regions",
        "AWS CloudTrail should be configured to log all management events across all AWS regions.",
        "high",
        "aws",
        "cloudtrail",
        "Logging",
        "(auto-calculated)",
        "Without CloudTrail, API activity is not recorded, making incident investigation impossible.",
        "Enable CloudTrail in all regions via AWS Console or CLI.",
        "cloudtrail_multi_region_enabled",
        "logging, monitoring, aws",
        "",
    ]
    for col_idx, val in enumerate(example, start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = Font(color="999999", italic=True)

    # ---- Sheet 2: Instructions ----
    ws_inst = wb.create_sheet("Instructions")
    instructions = [
        "CUSTOM CONTROLS IMPORT TEMPLATE",
        "",
        "Required fields: check_id, title, severity, provider",
        "",
        "Severity values: critical, high, medium, low, informational",
        "",
        "Provider values: aws, azure, gcp, oci, alibaba, ibm_cloud, kubernetes,",
        "  m365, github, google_workspace, salesforce, servicenow, snowflake, cloudflare, openstack",
        "",
        "Category values: Identity, Encryption, Storage, Networking, Logging, Compute,",
        "  Database, Container, Serverless, Data Protection, Backup, Compliance,",
        "  Threat Detection, Governance, Email Security, Collaboration, DevOps,",
        "  API Security, CDN, DNS, Analytics",
        "",
        "Assessment type is AUTO-DETERMINED:",
        "  - If mapped_scanner_check_ids contains valid IDs from the registry -> 'automated'",
        "  - If empty or no valid IDs -> 'manual'",
        "",
        "mapped_scanner_check_ids: Comma-separated scanner check IDs.",
        "  See the 'Available_Scanner_IDs' sheet for valid values.",
        "",
        "tags: Comma-separated. Use for organization and filtering.",
        "",
        "The row in the Controls sheet with italic text is an example - replace it with your data.",
    ]
    for i, line in enumerate(instructions, start=1):
        cell = ws_inst.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=12)
    ws_inst.column_dimensions["A"].width = 90

    # ---- Sheet 3: Available Scanner IDs ----
    ws_ref = wb.create_sheet("Available_Scanner_IDs")
    ref_headers = ["scanner_check_id", "provider", "title", "category", "assessment_type"]
    for col_idx, h in enumerate(ref_headers, start=1):
        cell = ws_ref.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    ws_ref.column_dimensions["A"].width = 45
    ws_ref.column_dimensions["B"].width = 15
    ws_ref.column_dimensions["C"].width = 60
    ws_ref.column_dimensions["D"].width = 20
    ws_ref.column_dimensions["E"].width = 15

    row = 2
    for check in sorted(registry.list_checks(), key=lambda c: c.check_id):
        for sid in check.scanner_check_ids or [check.check_id]:
            ws_ref.cell(row, 1, sid)
            ws_ref.cell(row, 2, check.provider)
            ws_ref.cell(row, 3, check.title)
            ws_ref.cell(row, 4, check.category)
            ws_ref.cell(row, 5, check.assessment_type)
            row += 1

    # ---- Sheet 4: Registry Checks ----
    ws_reg = wb.create_sheet("Registry_Checks")
    reg_headers = ["check_id", "title", "provider", "service", "category", "severity", "source", "cis_id"]
    for col_idx, h in enumerate(reg_headers, start=1):
        cell = ws_reg.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    ws_reg.column_dimensions["A"].width = 45
    ws_reg.column_dimensions["B"].width = 60
    ws_reg.column_dimensions["C"].width = 15
    ws_reg.column_dimensions["D"].width = 20
    ws_reg.column_dimensions["E"].width = 20
    ws_reg.column_dimensions["F"].width = 15
    ws_reg.column_dimensions["G"].width = 12
    ws_reg.column_dimensions["H"].width = 12

    row = 2
    for check in sorted(registry.list_checks(), key=lambda c: c.check_id):
        ws_reg.cell(row, 1, check.check_id)
        ws_reg.cell(row, 2, check.title)
        ws_reg.cell(row, 3, check.provider)
        ws_reg.cell(row, 4, check.service)
        ws_reg.cell(row, 5, check.category)
        ws_reg.cell(row, 6, check.severity)
        ws_reg.cell(row, 7, check.source)
        ws_reg.cell(row, 8, check.cis_id or "")
        row += 1

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
